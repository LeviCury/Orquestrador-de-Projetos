from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from datetime import date
from typing import Optional
import io

from ..database import get_db
from ..models import TimeEntry, TicketHourEntry, Project, Stage, Task, Collaborator
from ..schemas import TimeEntryCreate, TimeEntryRead, TimeEntryUpdate, QuickTimeEntry

router = APIRouter()


@router.get("/", response_model=list[TimeEntryRead])
def list_time_entries(
    project_id: Optional[int] = None,
    collaborator_id: Optional[int] = None,
    stage_id: Optional[int] = None,
    task_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    q = db.query(TimeEntry).options(joinedload(TimeEntry.collaborator))
    if project_id:
        q = q.filter(TimeEntry.project_id == project_id)
    if collaborator_id:
        q = q.filter(TimeEntry.collaborator_id == collaborator_id)
    if stage_id:
        q = q.filter(TimeEntry.stage_id == stage_id)
    if task_id:
        q = q.filter(TimeEntry.task_id == task_id)
    if date_from:
        q = q.filter(TimeEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(TimeEntry.entry_date <= date_to)
    return q.order_by(TimeEntry.entry_date.desc(), TimeEntry.id.desc()).all()


@router.post("/", response_model=TimeEntryRead, status_code=201)
def create_time_entry(data: TimeEntryCreate, db: Session = Depends(get_db)):
    collab = db.query(Collaborator).filter(Collaborator.id == data.collaborator_id).first()
    if not collab:
        raise HTTPException(404, "Collaborator not found")
    if collab.system_role != "analyst":
        raise HTTPException(403, "Apenas analistas podem registrar horas de projeto.")
    if not db.query(Project).filter(Project.id == data.project_id).first():
        raise HTTPException(404, "Project not found")
    if data.stage_id and not db.query(Stage).filter(Stage.id == data.stage_id).first():
        raise HTTPException(404, "Stage not found")
    if data.task_id and not db.query(Task).filter(Task.id == data.task_id).first():
        raise HTTPException(404, "Task not found")

    entry = TimeEntry(**data.model_dump())
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return db.query(TimeEntry).options(joinedload(TimeEntry.collaborator)).filter(
        TimeEntry.id == entry.id
    ).first()


@router.post("/quick", response_model=TimeEntryRead, status_code=201)
def quick_time_entry(data: QuickTimeEntry, db: Session = Depends(get_db)):
    """Simplified time entry -- auto-fills today's date."""
    collab = db.query(Collaborator).filter(Collaborator.id == data.collaborator_id).first()
    if not collab:
        raise HTTPException(404, "Collaborator not found")
    if collab.system_role != "analyst":
        raise HTTPException(403, "Apenas analistas podem registrar horas de projeto.")
    if not db.query(Project).filter(Project.id == data.project_id).first():
        raise HTTPException(404, "Project not found")

    entry_date_val = date.today()
    if data.entry_date:
        try:
            entry_date_val = date.fromisoformat(data.entry_date)
        except ValueError:
            pass

    entry = TimeEntry(
        collaborator_id=data.collaborator_id,
        project_id=data.project_id,
        stage_id=data.stage_id,
        task_id=data.task_id,
        entry_date=entry_date_val,
        hours_worked=data.hours_worked,
        description=data.description,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return db.query(TimeEntry).options(joinedload(TimeEntry.collaborator)).filter(
        TimeEntry.id == entry.id
    ).first()


@router.put("/{entry_id}", response_model=TimeEntryRead)
def update_time_entry(entry_id: int, data: TimeEntryUpdate, db: Session = Depends(get_db)):
    entry = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Time entry not found")
    updates = data.model_dump(exclude_unset=True)
    for key, val in updates.items():
        setattr(entry, key, val)
    db.commit()
    db.refresh(entry)
    return db.query(TimeEntry).options(joinedload(TimeEntry.collaborator)).filter(
        TimeEntry.id == entry.id
    ).first()


@router.delete("/{entry_id}", status_code=204)
def delete_time_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Time entry not found")
    db.delete(entry)
    db.commit()


@router.get("/summary/by-collaborator")
def hours_by_collaborator(project_id: int, db: Session = Depends(get_db)):
    rows = (
        db.query(
            Collaborator.id,
            Collaborator.name,
            func.coalesce(func.sum(TimeEntry.hours_worked), 0).label("total_hours"),
        )
        .join(TimeEntry, TimeEntry.collaborator_id == Collaborator.id)
        .filter(TimeEntry.project_id == project_id, Collaborator.system_role == "analyst")
        .group_by(Collaborator.id, Collaborator.name)
        .all()
    )
    return [{"collaborator_id": r[0], "name": r[1], "total_hours": float(r[2])} for r in rows]


@router.get("/unified")
def unified_hours(
    collaborator_id: Optional[int] = None,
    project_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Returns project hours + ticket hours in a single unified list."""
    pq = db.query(TimeEntry).options(joinedload(TimeEntry.collaborator))
    tq = db.query(TicketHourEntry)
    if collaborator_id:
        pq = pq.filter(TimeEntry.collaborator_id == collaborator_id)
        tq = tq.filter(TicketHourEntry.collaborator_id == collaborator_id)
    if project_id:
        pq = pq.filter(TimeEntry.project_id == project_id)
    if date_from:
        pq = pq.filter(TimeEntry.entry_date >= date_from)
        tq = tq.filter(TicketHourEntry.entry_date >= date_from)
    if date_to:
        pq = pq.filter(TimeEntry.entry_date <= date_to)
        tq = tq.filter(TicketHourEntry.entry_date <= date_to)

    project_names = {p.id: p.name for p in db.query(Project.id, Project.name).all()}
    stage_names = {s.id: s.name for s in db.query(Stage.id, Stage.name).all()}
    task_names = {t.id: t.name for t in db.query(Task.id, Task.name).all()}

    items = []
    for e in pq.order_by(TimeEntry.entry_date.desc(), TimeEntry.id.desc()).all():
        items.append({
            "id": e.id,
            "source": "project",
            "collaborator_id": e.collaborator_id,
            "collaborator_name": e.collaborator.name if e.collaborator else "",
            "entry_date": e.entry_date.isoformat() if e.entry_date else "",
            "hours_worked": float(e.hours_worked),
            "description": e.description or "",
            "project_id": e.project_id,
            "project_name": project_names.get(e.project_id, ""),
            "stage_name": stage_names.get(e.stage_id) if e.stage_id else None,
            "task_name": task_names.get(e.task_id) if e.task_id else None,
            "glpi_ticket_id": None,
            "glpi_ticket_title": None,
            "glpi_link": None,
            "created_at": e.created_at.isoformat() if e.created_at else "",
        })

    for e in tq.order_by(TicketHourEntry.entry_date.desc(), TicketHourEntry.created_at.desc()).all():
        collab = e.collaborator
        items.append({
            "id": e.id,
            "source": "ticket",
            "collaborator_id": e.collaborator_id,
            "collaborator_name": collab.name if collab else "",
            "entry_date": e.entry_date.isoformat() if e.entry_date else "",
            "hours_worked": float(e.hours_worked),
            "description": e.glpi_ticket_title or "",
            "project_id": None,
            "project_name": None,
            "stage_name": None,
            "task_name": None,
            "glpi_ticket_id": e.glpi_ticket_id,
            "glpi_ticket_title": e.glpi_ticket_title,
            "glpi_link": e.glpi_link,
            "glpi_status": getattr(e, "glpi_status", "") or "",
            "glpi_type": getattr(e, "glpi_type", "") or "",
            "glpi_priority": getattr(e, "glpi_priority", "") or "",
            "glpi_open_date": getattr(e, "glpi_open_date", "") or "",
            "glpi_assigned_to": getattr(e, "glpi_assigned_to", "") or "",
            "created_at": e.created_at.isoformat() if e.created_at else "",
        })

    items.sort(key=lambda x: (x["entry_date"], x["created_at"]), reverse=True)
    return items


@router.get("/export-excel")
def export_excel(
    collaborator_id: Optional[int] = None,
    project_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Generates a polished Excel workbook with all hours (project + ticket)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    rows_data = unified_hours(collaborator_id, project_id, date_from, date_to, db)

    wb = Workbook()

    # ─── Colors ───
    MINERVA_DARK = "1A3550"
    MINERVA_MID = "4A7FA5"
    MINERVA_GOLD = "C7B475"
    MINERVA_RED = "E83948"
    WHITE = "FFFFFF"
    LIGHT_BG = "F1F5F9"
    BORDER_CLR = "CBD5E1"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_CLR),
        right=Side(style="thin", color=BORDER_CLR),
        top=Side(style="thin", color=BORDER_CLR),
        bottom=Side(style="thin", color=BORDER_CLR),
    )

    # ─── Sheet 1: Detalhamento ───
    ws = wb.active
    ws.title = "Detalhamento de Horas"
    ws.sheet_properties.tabColor = MINERVA_DARK

    headers = [
        "Data", "Colaborador", "Tipo", "Projeto / Chamado", "Etapa", "Tarefa",
        "Horas", "Descrição", "Status GLPI", "Tipo GLPI", "Prioridade", "Técnico GLPI",
    ]
    col_widths = [14, 28, 14, 32, 22, 22, 10, 50, 20, 16, 14, 28]

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Relatório de Horas — Minerva Foods"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=MINERVA_DARK, end_color=MINERVA_DARK, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    # Subtitle
    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    period_from = date_from.isoformat() if date_from else "início"
    period_to = date_to.isoformat() if date_to else "hoje"
    sub_cell.value = f"Período: {period_from} a {period_to}  •  {len(rows_data)} lançamento(s)"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color=MINERVA_MID)
    sub_cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # Headers (row 4)
    header_fill = PatternFill(start_color=MINERVA_MID, end_color=MINERVA_MID, fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[4].height = 30

    # Data rows
    project_fill = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
    ticket_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    hours_font_project = Font(name="Calibri", size=10, bold=True, color=MINERVA_MID)
    hours_font_ticket = Font(name="Calibri", size=10, bold=True, color="D97706")

    total_project = 0.0
    total_ticket = 0.0

    for i, row in enumerate(rows_data):
        r = 5 + i
        is_project = row["source"] == "project"
        fill = project_fill if is_project else ticket_fill

        tipo = "Projeto" if is_project else "Chamado"
        ref = row["project_name"] or ""
        if not is_project and row.get("glpi_ticket_id"):
            ref = f"#{row['glpi_ticket_id']} — {row.get('glpi_ticket_title', '')}"

        vals = [
            row["entry_date"],
            row["collaborator_name"],
            tipo,
            ref,
            row.get("stage_name") or "",
            row.get("task_name") or "",
            row["hours_worked"],
            row["description"],
            row.get("glpi_status", "") if not is_project else "",
            row.get("glpi_type", "") if not is_project else "",
            row.get("glpi_priority", "") if not is_project else "",
            row.get("glpi_assigned_to", "") if not is_project else "",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 8))
            if col_idx == 7:
                cell.font = hours_font_project if is_project else hours_font_ticket
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 1:
                cell.number_format = "DD/MM/YYYY"
            if col_idx in (3, 9, 10, 11):
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if is_project:
            total_project += row["hours_worked"]
        else:
            total_ticket += row["hours_worked"]

    # Totals row
    total_row = 5 + len(rows_data) + 1
    total_fill = PatternFill(start_color=MINERVA_DARK, end_color=MINERVA_DARK, fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(row=total_row, column=1, value="TOTAL GERAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = thin_border
    total_cell = ws.cell(row=total_row, column=7, value=total_project + total_ticket)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = "0.0"
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.border = thin_border
    for c in range(8, 13):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = thin_border
    ws.row_dimensions[total_row].height = 30

    # Breakdown row
    brk = total_row + 1
    gold_fill = PatternFill(start_color=MINERVA_GOLD, end_color=MINERVA_GOLD, fill_type="solid")
    brk_font = Font(name="Calibri", size=10, bold=True, color=MINERVA_DARK)
    ws.merge_cells(start_row=brk, start_column=1, end_row=brk, end_column=6)
    ws.cell(row=brk, column=1, value=f"Projetos: {total_project:.1f}h   |   Chamados: {total_ticket:.1f}h").font = brk_font
    ws.cell(row=brk, column=1).fill = gold_fill
    ws.cell(row=brk, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 13):
        ws.cell(row=brk, column=c).fill = gold_fill
        ws.cell(row=brk, column=c).border = thin_border
    ws.row_dimensions[brk].height = 26

    # ─── Sheet 2: Resumo por Colaborador ───
    ws2 = wb.create_sheet("Resumo por Colaborador")
    ws2.sheet_properties.tabColor = MINERVA_GOLD

    by_collab: dict[str, dict] = defaultdict(lambda: {"project": 0.0, "ticket": 0.0, "days": set()})
    for row in rows_data:
        name = row["collaborator_name"]
        by_collab[name][row["source"]] += row["hours_worked"]
        by_collab[name]["days"].add(row["entry_date"])

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = "Resumo por Colaborador"
    t2.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t2.fill = PatternFill(start_color=MINERVA_DARK, end_color=MINERVA_DARK, fill_type="solid")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 38

    h2_headers = ["Colaborador", "Horas Projetos", "Horas Chamados", "Total", "Dias Trabalhados", "Média/Dia"]
    h2_widths = [30, 18, 18, 14, 20, 14]
    for col_idx, (h, w) in enumerate(zip(h2_headers, h2_widths), 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[3].height = 28

    sorted_collabs = sorted(by_collab.items(), key=lambda x: x[1]["project"] + x[1]["ticket"], reverse=True)
    for i, (name, data) in enumerate(sorted_collabs):
        r = 4 + i
        total = data["project"] + data["ticket"]
        days = len(data["days"])
        avg = total / days if days > 0 else 0
        fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid") if i % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        vals = [name, round(data["project"], 1), round(data["ticket"], 1), round(total, 1), days, round(avg, 1)]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.font = data_font if col_idx != 4 else Font(name="Calibri", size=10, bold=True, color=MINERVA_DARK)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            if col_idx in (2, 3, 4, 6):
                cell.number_format = "0.0"

    # ─── Sheet 3: Resumo por Data ───
    ws3 = wb.create_sheet("Resumo por Data")
    ws3.sheet_properties.tabColor = MINERVA_RED

    by_date: dict[str, dict] = defaultdict(lambda: {"project": 0.0, "ticket": 0.0})
    for row in rows_data:
        by_date[row["entry_date"]][row["source"]] += row["hours_worked"]

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = "Resumo por Data"
    t3.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t3.fill = PatternFill(start_color=MINERVA_DARK, end_color=MINERVA_DARK, fill_type="solid")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 38

    h3_headers = ["Data", "Horas Projetos", "Horas Chamados", "Total", "Meta (9h)"]
    h3_widths = [16, 18, 18, 14, 14]
    for col_idx, (h, w) in enumerate(zip(h3_headers, h3_widths), 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.row_dimensions[3].height = 28

    for i, (dt, data) in enumerate(sorted(by_date.items(), reverse=True)):
        r = 4 + i
        total = data["project"] + data["ticket"]
        status = "OK" if total >= 9 else f"Faltam {9 - total:.1f}h"
        fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid") if i % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        vals = [dt, round(data["project"], 1), round(data["ticket"], 1), round(total, 1), status]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col_idx, value=val)
            cell.font = data_font if col_idx != 4 else Font(name="Calibri", size=10, bold=True)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (2, 3, 4):
                cell.number_format = "0.0"
            if col_idx == 5:
                if total >= 9:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="059669")
                else:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=MINERVA_RED)

    # ─── Sheet 4: Resumo por Chamado GLPI ───
    ws4 = wb.create_sheet("Chamados GLPI")
    ws4.sheet_properties.tabColor = "D97706"

    ticket_rows = [r for r in rows_data if r["source"] == "ticket"]
    by_ticket: dict[str, dict] = {}
    for row in ticket_rows:
        tid = row.get("glpi_ticket_id") or "sem-id"
        if tid not in by_ticket:
            by_ticket[tid] = {
                "title": row.get("glpi_ticket_title", ""),
                "status": row.get("glpi_status", ""),
                "type": row.get("glpi_type", ""),
                "priority": row.get("glpi_priority", ""),
                "assigned_to": row.get("glpi_assigned_to", ""),
                "link": row.get("glpi_link", ""),
                "total_hours": 0.0,
                "entries": 0,
                "collaborators": set(),
                "dates": set(),
            }
        by_ticket[tid]["total_hours"] += row["hours_worked"]
        by_ticket[tid]["entries"] += 1
        by_ticket[tid]["collaborators"].add(row["collaborator_name"])
        by_ticket[tid]["dates"].add(row["entry_date"])

    ws4.merge_cells("A1:J1")
    t4 = ws4["A1"]
    t4.value = "Resumo por Chamado GLPI"
    t4.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t4.fill = PatternFill(start_color=MINERVA_DARK, end_color=MINERVA_DARK, fill_type="solid")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 38

    h4_headers = [
        "Chamado", "Título", "Status", "Tipo", "Prioridade",
        "Técnico GLPI", "Lançado por", "Horas Total", "Lançamentos", "Link",
    ]
    h4_widths = [16, 40, 22, 16, 14, 28, 28, 14, 14, 50]
    for col_idx, (h, w) in enumerate(zip(h4_headers, h4_widths), 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws4.column_dimensions[get_column_letter(col_idx)].width = w
    ws4.row_dimensions[3].height = 28

    sorted_tickets = sorted(by_ticket.items(), key=lambda x: x[1]["total_hours"], reverse=True)
    for i, (tid, data) in enumerate(sorted_tickets):
        r = 4 + i
        fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid") if i % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        vals = [
            f"#{tid}" if tid != "sem-id" else "",
            data["title"],
            data["status"],
            data["type"],
            data["priority"],
            data["assigned_to"],
            ", ".join(sorted(data["collaborators"])),
            round(data["total_hours"], 1),
            data["entries"],
            data["link"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=r, column=col_idx, value=val)
            cell.font = data_font if col_idx != 8 else Font(name="Calibri", size=10, bold=True, color="D97706")
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3, 4, 5, 8, 9) else "left",
                vertical="center",
                wrap_text=(col_idx in (2, 7, 10)),
            )
            if col_idx == 8:
                cell.number_format = "0.0"

    # Total row for tickets sheet
    if sorted_tickets:
        t4_total_row = 4 + len(sorted_tickets) + 1
        ws4.merge_cells(start_row=t4_total_row, start_column=1, end_row=t4_total_row, end_column=7)
        ws4.cell(row=t4_total_row, column=1, value="TOTAL CHAMADOS").font = total_font
        ws4.cell(row=t4_total_row, column=1).fill = total_fill
        ws4.cell(row=t4_total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        for c in range(1, 11):
            ws4.cell(row=t4_total_row, column=c).fill = total_fill
            ws4.cell(row=t4_total_row, column=c).border = thin_border
        t4_total_cell = ws4.cell(row=t4_total_row, column=8, value=total_ticket)
        t4_total_cell.font = total_font
        t4_total_cell.fill = total_fill
        t4_total_cell.number_format = "0.0"
        t4_total_cell.alignment = Alignment(horizontal="center", vertical="center")
        t4_entries_cell = ws4.cell(row=t4_total_row, column=9, value=len(ticket_rows))
        t4_entries_cell.font = total_font
        t4_entries_cell.fill = total_fill
        t4_entries_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws4.row_dimensions[t4_total_row].height = 30

    # ─── Freeze panes & print setup ───
    ws.freeze_panes = "A5"
    ws2.freeze_panes = "A4"
    ws3.freeze_panes = "A4"
    ws4.freeze_panes = "A4"
    ws.print_area = f"A1:L{brk}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"horas_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
